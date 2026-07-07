from __future__ import annotations

import json
import re
import shutil
from collections import Counter
from pathlib import Path

import openpyxl
from docx import Document


ROOT = Path(r"C:\Users\Administrator\Documents\exam")
APP_FILE = ROOT / "index.html"
REPORT_FILE = ROOT / "新题导入报告.json"
BACKUP_FILE = ROOT / "index_新增ExcelDocx题目前备份.html"

XLSX_FILES = {
    "IA": Path(r"C:\Users\Administrator\Desktop\信息架构参考(1).xlsx"),
    "BA": Path(r"C:\Users\Administrator\Desktop\业务架构参考(1).xlsx"),
    "AA": Path(r"C:\Users\Administrator\Desktop\应用架构参考(1).xlsx"),
}
DOCX_FILE = Path(r"C:\Users\Administrator\Desktop\知识点\02试题整理\单选多选.docx")

TYPE_MAP = {
    "单选": "single",
    "单选题": "single",
    "单项选择题": "single",
    "多选": "multiple",
    "多选题": "multiple",
    "多项选择题": "multiple",
    "判断": "single",
    "判断题": "single",
}

DOMAIN_HINTS = {
    "IA": [
        "信息架构", "数据架构", "逻辑实体", "属性", "主标识符", "第二范式", "第三范式",
        "数据实体", "数据模型", "数据标准", "主数据", "事务数据", "编码规则", "数据项",
    ],
    "BA": [
        "业务架构", "业务单元", "业务活动", "业务流程", "能力流程", "业务组件",
        "价值流", "业务能力", "业务场景", "场景因子", "灯塔指标", "流程分类", "数字孪生",
    ],
    "AA": [
        "应用架构", "应用组件", "组件契约", "应用服务", "应用上下文", "应用交互",
        "时序图", "接口", "API", "微服务", "应用资产", "应用域", "应用系统", "技术架构",
    ],
}


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def normalize_key(text: str) -> str:
    text = re.sub(r"^\s*\d+\s*[.、]\s*", "", text)
    text = re.sub(r"【[^】]+】", "", text)
    return re.sub(r"[\W_]+", "", text).lower()


def read_app_questions() -> tuple[str, list[dict]]:
    html = APP_FILE.read_text(encoding="utf-8-sig")
    match = re.search(r"const questions = (\[.*?\]);", html, re.S)
    if not match:
        raise RuntimeError("index.html 中没有找到 const questions 数组")
    return html, json.loads(match.group(1))


def replace_app_questions(html: str, questions: list[dict]) -> str:
    js = "    const questions = " + json.dumps(questions, ensure_ascii=False, indent=6) + ";"
    js = js.replace("\n", "\n    ", 1)
    marker = "    const questions = "
    start = html.find(marker)
    if start < 0:
        raise RuntimeError("???????????? const questions")
    array_start = html.find("[", start)
    if array_start < 0:
        raise RuntimeError("????????????????")
    depth = 0
    in_string = False
    escape = False
    end = -1
    for idx in range(array_start, len(html)):
        ch = html[idx]
        if in_string:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_string = False
        else:
            if ch == '"':
                in_string = True
            elif ch == "[":
                depth += 1
            elif ch == "]":
                depth -= 1
                if depth == 0:
                    semi = html.find(";", idx)
                    if semi < 0:
                        raise RuntimeError("??????????????")
                    end = semi + 1
                    break
    if end < 0:
        raise RuntimeError("??????????????")
    return html[:start] + js + html[end:]

def parse_answer(answer: str, qtype: str) -> list[str]:
    answer = clean(answer).upper().replace("，", "").replace(",", "").replace("、", "").replace(" ", "")
    if answer in {"正确", "√", "TRUE", "T"}:
        return ["A"]
    if answer in {"错误", "×", "FALSE", "F"}:
        return ["B"]
    letters = re.findall(r"[A-H]", answer)
    if qtype == "single" and len(letters) > 1:
        return letters[:1]
    return letters


def parse_xlsx(path: Path, business: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[0]
    header = [clean(ws.cell(2, c).value) for c in range(1, ws.max_column + 1)]

    def find_col(*needles: str) -> int | None:
        for idx, name in enumerate(header):
            compact = re.sub(r"\s+", "", name)
            if any(needle in compact for needle in needles):
                return idx + 1
        return None

    stem_col = find_col("题干")
    type_col = find_col("题型")
    answer_col = find_col("正确答案", "答案")
    analysis_col = find_col("解析")
    if not stem_col or not type_col or not answer_col:
        raise RuntimeError(f"{path} 表头缺少题干/题型/正确答案列")

    option_cols = []
    for letter in "ABCDEFGH":
        col = find_col(f"选项{letter}", f"选项 {letter}")
        if col:
            option_cols.append((letter, col))

    items = []
    for row in range(3, ws.max_row + 1):
        stem = clean(ws.cell(row, stem_col).value)
        raw_type = clean(ws.cell(row, type_col).value)
        qtype = TYPE_MAP.get(raw_type)
        if not stem or qtype not in {"single", "multiple"}:
            continue
        options = [clean(ws.cell(row, col).value) for _, col in option_cols]
        options = [item for item in options if item]
        if raw_type in {"判断", "判断题"} and not options:
            options = ["正确", "错误"]
        answer = parse_answer(clean(ws.cell(row, answer_col).value), qtype)
        if not options or not answer:
            continue
        analysis = clean(ws.cell(row, analysis_col).value) if analysis_col else ""
        items.append({
            "type": qtype,
            "business": business,
            "question": stem,
            "options": options,
            "answer": answer,
            "analysis": analysis or f"本题答案为{''.join(answer)}。",
            "source": path.name,
        })
    return items


def split_docx_blocks(path: Path) -> list[str]:
    doc = Document(str(path))
    lines = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    blocks = []
    current = []
    for line in lines:
        if re.match(r"^(?:\d+\s*[.、]\s*)?【(?:单选|多选|单选题|多选题|判断|判断题)】", line):
            if current:
                blocks.append(current)
            current = [line]
        elif current:
            current.append(line)
    if current:
        blocks.append(current)
    return ["\n".join(block) for block in blocks]


def parse_docx_block(block: str) -> dict | None:
    lines = [line.strip() for line in block.splitlines() if line.strip()]
    if not lines:
        return None
    head = lines[0]
    type_match = re.search(r"【([^】]+)】", head)
    qtype = TYPE_MAP.get(type_match.group(1), "single") if type_match else "single"
    question = re.sub(r"^\s*\d+\s*[.、]\s*", "", head)
    question = re.sub(r"【[^】]+】", "", question).strip()
    options = {}
    answer = ""
    analyses = []
    current_option = None
    mode = "question"
    for line in lines[1:]:
        option_match = re.match(r"^([A-H])[.、]\s*(.+)$", line)
        answer_match = re.match(r"^答案[:：]\s*(.+)$", line)
        analysis_match = re.match(r"^解析[:：]\s*(.*)$", line)
        if option_match:
            current_option = option_match.group(1)
            options[current_option] = option_match.group(2).strip()
            mode = "options"
        elif answer_match:
            answer = answer_match.group(1).strip()
            current_option = None
            mode = "answer"
        elif analysis_match:
            text = analysis_match.group(1).strip()
            if text:
                analyses.append(text)
            current_option = None
            mode = "analysis"
        elif mode == "analysis":
            analyses.append(line)
        elif current_option:
            options[current_option] = f"{options[current_option]} {line}".strip()
        else:
            question = f"{question} {line}".strip()
    if qtype not in {"single", "multiple"}:
        return None
    option_values = [options[key] for key in sorted(options)]
    parsed_answer = parse_answer(answer, qtype)
    if not option_values or not parsed_answer:
        return None
    return {
        "type": qtype,
        "business": "",
        "question": question,
        "options": option_values,
        "answer": parsed_answer,
        "analysis": " ".join(analyses).strip() or f"本题答案为{''.join(parsed_answer)}。",
        "source": DOCX_FILE.name,
    }


def score_business(item: dict) -> tuple[str, dict[str, int]]:
    text = f"{item['question']} {' '.join(item.get('options', []))} {item.get('analysis', '')}"
    scores = {key: 0 for key in ("IA", "BA", "AA")}
    for business, hints in DOMAIN_HINTS.items():
        for hint in hints:
            if hint.lower() in text.lower():
                scores[business] += 10
    best = max(scores, key=scores.get)
    if scores[best] == 0:
        best = "BA"
    return best, scores


def renumber(items: list[dict]) -> list[dict]:
    for idx, item in enumerate(items, 1):
        item["id"] = f"q{idx:04d}"
    return items


def main() -> None:
    html, existing = read_app_questions()
    existing_keys = {normalize_key(item["question"]) for item in existing}
    candidates = []
    source_counts = Counter()

    for business, path in XLSX_FILES.items():
        items = parse_xlsx(path, business)
        candidates.extend(items)
        source_counts[f"xlsx:{business}"] += len(items)

    docx_classification = []
    for block in split_docx_blocks(DOCX_FILE):
        item = parse_docx_block(block)
        if not item:
            continue
        business, scores = score_business(item)
        item["business"] = business
        candidates.append(item)
        source_counts["docx"] += 1
        docx_classification.append({"question": item["question"], "business": business, "scores": scores})

    imported = []
    skipped = []
    seen_new = set()
    for item in candidates:
        key = normalize_key(item["question"])
        if not key:
            continue
        if key in existing_keys or key in seen_new:
            skipped.append({"question": item["question"], "business": item["business"], "source": item["source"]})
            continue
        seen_new.add(key)
        imported.append({
            "id": "",
            "type": item["type"],
            "business": item["business"],
            "question": item["question"],
            "options": item["options"],
            "answer": item["answer"],
            "analysis": item["analysis"],
        })

    updated = renumber(existing + imported)
    if not BACKUP_FILE.exists():
        shutil.copy2(APP_FILE, BACKUP_FILE)
    APP_FILE.write_text(replace_app_questions(html, updated), encoding="utf-8")

    report = {
        "existing_count": len(existing),
        "candidate_count": len(candidates),
        "imported_count": len(imported),
        "skipped_duplicate_count": len(skipped),
        "updated_count": len(updated),
        "source_counts": dict(source_counts),
        "imported_by_business": dict(Counter(item["business"] for item in imported)),
        "imported_by_type": dict(Counter(item["type"] for item in imported)),
        "docx_classification": docx_classification,
        "skipped_duplicates_sample": skipped[:80],
        "backup": str(BACKUP_FILE),
    }
    REPORT_FILE.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
